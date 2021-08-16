"""
Various classes and methods for inputting formatted XML files.

After files have been processed using the ParsedXML module, their relevant
attributes are directly accessible. From there, it is easy to automate
their entry into the relevant spreadsheets.

For ease of use, this module assumes responsbility for the following:
    - Creating new spreadsheets and worksheets as necessary.
    - Reading the ENERGYFIN directory structure into file lists.
    - Converting lists of files to ParsedXML files.
    - Formatting and modifying spreadsheets for data entry. 

The SheetHandler module is an abstract base class and cannot be instantiated
directly. Instead, SheetHandler.get_handler takes a subcommand aka parsetype,
either 'invoice' or 'summary', along with an output directory and template
file, and returns a concrete subclass.

The file typing hierarchy is as follows:
    invoice [class]
        \- miso
            \- mkt {xml} + ca {xml}
        \- pjm
            \- {xml}
    summary [class]
        \- miso
            \- {zip}
                \- ao {xml} + ftr_S7 {xml}
"""

import pickle
from abc import ABC, abstractmethod, abstractstaticmethod
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, TypeVar

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from parsers import Invoice, Settlement


S = TypeVar(name='S', bound='SheetHandler')
"""type hint for sheethandler class before definition"""


class SheetHandler(ABC):
    """
    Abstract class for excel handling.

    This class dictates the structure of all SheetHandlers using abstract
    functions, and defines multiple shared helper functions.

    It also lays the groundwork for the file-sheet interface formalized in
    its subclasses.
    """
    def __init__(self, path: Path, template: Workbook) -> None:
        self.path = path
        self.path.mkdir(parents=True, exist_ok=True)
        """The directory to contain all related workbooks."""

        self.template = template
        """Workbook with named template worksheets."""

        self.modified: Dict[Path, Workbook] = {}
        """An index for repeated access to an already-loaded spreadsheet."""
    
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        """The active workbook and worksheet."""

    @staticmethod
    def get_handler(parsetype: str, path: Path, template: Path) -> S:
        """
        Factory function.

        Loads the template archive; then depending on whether invoices or
        settlements are being parsed, returns the correct <dict> map.

        Returns an instance of the correct SheetHandler.
        """
        with open(template, 'rb') as file:
            template = pickle.load(file)[parsetype]
        if parsetype == 'invoice':
            return InvoiceHandler(path, template)
        elif parsetype == 'settlement':
            return SettlementHandler(path, template)

    @staticmethod
    def _get_adjacent_years(date: datetime) -> List[int]:
        """
        If the month is January or December, it should also be filled in the
        spreadsheet of the relevant neighboring year.
        """
        years = [date.year]
        if date.month == 1:
            years.append(date.year - 1)
        elif date.month == 12:
            years.append(date.year + 1)
        return years

    def _get_template(self, key: str) -> Worksheet:
        return self.template[key]

    def _set_workbook(self, name) -> None:
        """
        Loads a workbook only if it hasn't been loaded yet.
        
        If it doesn't exist, creates a new workbook and removes the default
        empty spreadsheet.
        """
        path = self.path / name
        if path not in self.modified:
            try:
                wb = load_workbook(path)
            except FileNotFoundError:
                wb = Workbook()
                wb.remove(wb.active)
            finally:
                self.modified[path] = wb
        self.workbook = self.modified[path]
        self.worksheet = None

    def _set_sheet(self, sheet_name: str) -> bool:
        """
        Changes the active worksheet if necessary.

        Creates a new worksheet if it doesn't exist.
        """
        if not (self.worksheet and self.worksheet.title == sheet_name):
            try:
                self.worksheet = self.workbook[sheet_name]
            except KeyError:
                self.worksheet = self.workbook.create_sheet(sheet_name)
                return False
            else:
                return True
    
    def _search(self, kwd: str, col: int) -> Optional[list]:
        """
        Returns the index of a string (by row) in a given column, or None.
        """
        for row in range(1, self.worksheet.max_row + 1):
            if self.worksheet.cell(row, col).value == kwd:
                return row
        return None

    def _copy_cell(self, cell: Cell, temp_cell: Cell) -> None:
        """
        Copies all cell attributes and pastes its contents, translating any
        formulas by relative position.
        """
        for attr in ['font', 'border', 'fill', 'number_format', 'alignment']:
            setattr(cell, attr, copy(getattr(temp_cell, attr)))
        translator = Translator(temp_cell.value, temp_cell.coordinate)
        cell.value = translator.translate_formula(cell.coordinate)

    def _set_cell(self, row: int, col: int, value: Any) -> None:
        """
        Sets a cell index to any value.
        """
        self.worksheet.cell(row, col).value = value

    def _fill_template(self, row: int, template: Worksheet) -> None:
        """
        Pastes a new section of the stored template at the specified
        row, relative to its position.
        """
        for col, dim in template.column_dimensions.items():
            self.worksheet.column_dimensions[col].width = dim.width
        for r in range(1, template.max_row + 1):
            for c in range(1, template.max_column + 1):
                v = template.cell(r, c)
                cell = self.worksheet.cell(r+(row-1), c)
                if not cell.value:
                    self._copy_cell(cell, v)

    def write(self) -> None:
        """
        Writes out all modified workbooks.
        """
        for path, workbook in self.modified.items():
            workbook.save(path)

    """
    The following methods must be implemented in any subclass of SheetHandler.
    """

    @abstractstaticmethod
    def _date_to_col(self) -> int:
        pass

    @abstractmethod
    def _paste(self) -> Optional[int]:
        pass

    @abstractmethod
    def _fill_column(self) -> None:
        pass

    @abstractmethod
    def process_dir(self, input, market, move) -> None:
        pass

class InvoiceHandler(SheetHandler):
    """
    Implementation of SheetHandler for invoice files.
    """
    @staticmethod
    def _date_to_col(date: datetime) -> int:
        """
        Invoices are received at weekly intervals; convert a datetime into
        the proper column number based on market specifics.

        An invoice is sent on the last day of the month, which requires an
        extra column.
        """
        week = (date.day - 1) // 7
        col = week + 2
        if date.month != (date + relativedelta(days=1)).month:
            col += 1  # Extra column for the last (incomplete) week
        return col

    def _get_month_row(self, date: datetime) -> Optional[int]:
        """
        Invoice entries are separated by month; search for the proper section
        based on a datetime.
        """
        month = date.strftime('%B')
        row = self._search(month, 2)
        return row

    def _paste(self, date: datetime, template: Worksheet) -> int:
        """
        Create a new section with the given month name below the lowest entry.

        XXX: doesn't automatically sort jan-dec
        """
        row = self.worksheet.max_row + int(self.worksheet.max_row > 1)
        self._fill_template(row, template)
        self._set_cell(row, 2, date.strftime('%B'))
        return row

    def _fill_column(self, head: int, col: int,
                     invoice: Invoice, market: str) -> None:
        """
        Given an Invoice, fills the revenue, date and fees respectively.
        """
        if market == 'miso':
            self._set_cell(head+2, col, invoice.revenue)
            self._set_cell(head+4, col, invoice.date)
            self._set_cell(head+5, col, invoice.fees)
        elif market == 'pjm':
            self._set_cell(head+2, col, invoice.date)
            for i, amt in enumerate(invoice.amts, start=3):
                self._set_cell(head+i, col, amt)

    def fill(self, invoice: Invoice, market: str) -> None:
        """
        Given a Invoice, fills the information for the given week in the
        correct spreadsheet, worksheet, row and column.

        MISO splits the last week of the month, so the last week of each month
        is also added to the last column of the prior month.
        """
        if market == 'miso':
            for year in self._get_adjacent_years(invoice.date):
                self._set_workbook(f'{year}.xlsx')
                self._set_sheet(invoice.fund)

                row = self._get_month_row(invoice.date)
                col = self._date_to_col(invoice.date)
                if not row:
                    row = self._paste(invoice.date, self._get_template(market))

                self._fill_column(row, col, invoice, market)

                last_month = invoice.date - relativedelta(months=1)
                prev_row = self._get_month_row(last_month)
                if col == 2 and prev_row:
                    """If the invoice is the first week of the month, fill it in
                    the previous month (if an entry exists)."""
                    self._fill_column(prev_row, col, invoice, market)

        elif market == 'pjm':
            for year in self._get_adjacent_years(invoice.date):
                self._set_workbook(f'{year}.xlsx')
                self._set_sheet(invoice.fund)

                row = self._get_month_row(invoice.date)
                col = self._date_to_col(invoice.date)
                if not row:
                    t = self._get_template(market)
                    template = self.template.copy_worksheet(t)
                    template.insert_rows(4, len(invoice.names) - 1)
                    temp_row = template[3 + len(invoice.names)]

                    for i, name in enumerate(invoice.names, start=4):
                        for c in range(1, template.max_column + 1):
                            self._copy_cell(template.cell(i, c), temp_row[c-1])
                        template.cell(i, 1).value = name
                    row = self._paste(invoice.date, template)

                self._fill_column(row, col, invoice, market)

    def process_dir(self, input: Path, market: str, move=False) -> None:
        """
        Given an input directory, process its files into Invoices, and then
        iteratively fill each one.
        """
        files = Invoice.from_dir(input, market, move)
        for file in files:
            self.fill(file, market)

class SettlementHandler(SheetHandler):
    @staticmethod
    def _date_to_col(date: datetime) -> int:
        return date.day + 1

    def _paste(self, template: Worksheet) -> None:
        self._fill_template(1, template)

    def _fill_column(self, col: int, amounts: Dict[str, float]) -> None:
        for name, val in amounts.items():
            row = self._search(name, 1)
            if not row:
                row = self.worksheet.max_row + 1
                self._set_cell(row, 1, name)
            self._set_cell(row, col, val)

    def fill_summary(self, settlement: Settlement) -> None:
        for year in self._get_adjacent_years(settlement.date):
            self._set_workbook(f'{settlement.fund} Summary {year}.xlsx')
            if not self._set_sheet(settlement.date.strftime('%B')):
                self._paste(self._get_template('summary'))

            col = self._date_to_col(settlement.date)

            self._fill_column(col, settlement.ao_amounts)

    def fill_ftr(self, settlement: Settlement) -> None:
        for year in self._get_adjacent_years(settlement.date):
            self._set_workbook(f'{settlement.fund} FTR {year}.xlsx')
            if not self._set_sheet(settlement.date.strftime('%B')):
                self._paste(self._get_template('ftr'))

            col = self._date_to_col(settlement.date)

            self._fill_column(col, settlement.ftr_amounts)

    def process_dir(self, input: Path, market: str, move=False) -> None:
        files = Settlement.from_dir(input, market, move)
        for file in files:
            self.fill_summary(file)
            self.fill_ftr(file)