import pickle
import smtplib

from argparse import ArgumentParser, Namespace
from copy import copy
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Callable, Optional
from xml.etree import ElementTree

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet


"""
Configuration.
"""

SCRIPT_DESC = "Takes a collection of invoices and sends to .xlsx"
SCRIPT_USAGE = "xl-parse.py [-h] <invoice|settlement> --output=DIR --input=DIR --template=FILE --doctype=<{ca,mkt}|{ao,ftr}> \n"

TO_EMAIL = 'venkat@oncept.net'
FROM_EMAIL = 'Me_watch@oncept.net'
SMTP_SERVER = '10.10.0.22'


"""
Core functions.
"""

def parse_args() -> Namespace:
    parser = ArgumentParser(description=SCRIPT_DESC, usage=SCRIPT_USAGE)

    subparsers = parser.add_subparsers(dest='parsetype', metavar="one of <invoice> or <settlement>", required=True)
    invoice = subparsers.add_parser('invoice')
    settlement = subparsers.add_parser('settlement')

    parser.add_argument('--output', default='.', metavar="DIR", type=Path, help="output directory")
    parser.add_argument('--input', default='.', metavar="DIR", type=Path, help="input directory")
    parser.add_argument('--template', default='template.pkl', metavar="FILE", type=Path, help="template file (preformatted)")
    # parser.add_argument('--log', const=".", metavar="DIR", nargs='?', type=Path, help="send a log to directory")
    # parser.add_argument('--market', choices=['miso', 'pjm'], default='miso', help="market of document")

    invoice.add_argument('--doctype', choices=['ca', 'mkt'], required=True, help="type of document")
    settlement.add_argument('--doctype', choices=['ao', 'ftr'], required=True, help="type of document")

    return parser.parse_args()

def reduce(doctype: str, ao: Any = 'ao', ftr: Any = 'ftr', invoice: Any = 'invoice') -> Any:
    if doctype == 'ao':
        return ao
    elif doctype == 'ftr':
        return ftr
    elif doctype in ['ca', 'mkt']:
        return invoice
    else:
        raise ValueError("Doctype not recognized")

def load_template(template: Path, key: str) -> dict:
    with open(template, 'rb') as file:
        template = pickle.load(file)[key]
    return template

def get_toname(doctype: str, output: Path) -> Callable[[int, str], Path]:
    ao = lambda year, fund: output / fund + ' AO ' + str(year) + '.xlsx'
    ftr = lambda year, fund: output / fund + ' FTR ' + str(year) + '.xlsx'
    invoice = lambda year, _: output / str(year) + '.xlsx'
    return reduce(doctype, ao, ftr, invoice)


"""
Email utils.
"""

def email(message, to_email: str=TO_EMAIL, subject: str="<no subject>",
          from_email: str=FROM_EMAIL, server: str=SMTP_SERVER) -> None:
    msg = EmailMessage()
    msg['To'] = ', '.join(to_email)
    msg['From'] = from_email
    msg['Subject'] = subject
    msg.set_content(message)

    server = smtplib.SMTP(server)
    server.send_message(msg)
    server.quit()


"""
XML parser class.
"""

class ParsedXML:
    paths = {
        'invoice': {
            'net_rev': 'Header/[Page_Num="1"]/Tot_Net_Chg_Rev_Amt',
            'end_date': 'Header/[Page_Num="1"]/Billing_Prd_End_Dte',
            'fund': 'Header/[Page_Num="1"]/Mrkt_Participant_NmAddr',
            'delta': 7
        },
        'ao': {

        },
        'ftr': {

        }
    }

    def __init__(self, file: Path, doctype: str) -> None:
        doctype = reduce(doctype)

        self.file = file
        paths = ParsedXML.paths[doctype]

        self.root = ElementTree.parse(self.file).getroot()

        self.net_rev = float(self.root.findtext(paths['net_rev']).replace(',', ''))
        self.fund = self.root.findtext(paths['fund'])
        self._end_date = self.root.findtext(paths['end_date'])
        self.date = datetime.strptime(self._end_date, '%m/%d/%Y') - relativedelta(days=paths['delta'])
        
        self.years = list(set((self.date + relativedelta(months=i)).year for i in range(-1, 2)))

    @classmethod
    def from_list(cls, input: list, doctype: str):
        return [cls(file, doctype) for file in input]


"""
Excel handlers.
"""

class SheetHandler:
    def __init__(self, path: Path, workbook: Workbook, template: Worksheet, forceful: bool=False) -> None:
        self.path = path
        self.workbook = workbook
        self.template = template
        self.forceful = forceful

        self.worksheet: Worksheet = self.workbook.active

    def _copy_cell(cell: Cell, temp_cell: Cell) -> None:
        for attr in ['font', 'border', 'fill', 'number_format', 'alignment']:
            setattr(cell, attr, copy(getattr(temp_cell, attr)))
        cell.value = Translator(temp_cell.value, temp_cell.coordinate).translate_formula(cell.coordinate)

    def _set_cell(self, row: int, col: int, value: Any) -> None:
        cell = self.worksheet.cell(row=row, column=col)
        if self.forceful or not cell.value:
            cell.value = value
 
    def _search(self, kwd: str, col: int) -> Optional[list]:
        for row in range(1, self.worksheet.max_row):
            if self.worksheet.cell(row=row, column=col).value == kwd:
                return row
        return None

    def _paste(self, row: int) -> None:
        row -= 1  # convert to offset
        for r in range(1, self.template.max_row + 1):
            for c in range(1, self.template.max_column + 1):
                v = self.template.cell(row=r, column=c)
                cell = self.worksheet.cell(row=r+row, column=c)

                if self.forceful or not cell.value:
                    self._copy_cell(cell, v)
    
    def _set_sheet(self, sheet_name: str) -> None:
        if not (self.worksheet and self.worksheet.title == sheet_name):
            try:
                self.worksheet = self.workbook[sheet_name]
            except KeyError:
                self.worksheet = self.workbook.create_sheet(sheet_name)

            for col, dim in self.template.column_dimensions.items():
                self.worksheet.column_dimensions[col].width = dim.width
 
    def write(self) -> None:
        self.workbook.save(self.path)

class InvoiceHandler(SheetHandler):
    def _fill_column(self, head: int, col: int, date: datetime, revenue: float) -> None:
        self._set_cell(head+2, col, revenue)
        self._set_cell(head+4, col, date)

    def fill(self, file: ParsedXML) -> None:
        self.set_sheet(file.fund)

        month = file.date.strftime('%B')
        last_month = (file.date - relativedelta(months=1)).strftime('%B')

        row = self._search(month, 2) 

        if not row:
            row = self.worksheet.max_row + 1
            self._paste(row)
            self.worksheet.cell(row=row, column=2).value = month

        week = (file.date.day - 1) // 7
        col = week + 2

        if file.date.month != (file.date + relativedelta(days=1)).month:
            col += 1

        self._fill_column(row, col, file.date, file.net_rev)

        if col == 2 and self._search_month(last_month):
            self._set_month(last_month)
            self._fill_column(row, 8, file.date, file.net_rev)

class AO(SheetHandler):
    def trade_results(self, day: int) -> float:
        ahead_energy_amt = self.worksheet.cell(row=3, column=day+1)
        real_energy_amt = self.worksheet.cell(row=4, column=day+1)
        return -sum(filter(None, [ahead_energy_amt, real_energy_amt]))
    
    def total(self, day: int) -> float:
        cols = [self.worksheet.cell(row=r, column=day+1).value for r in range(5, self.template.max_column + 1)]
        return self.trade_results - sum(filter(None, cols))

class FTR(SheetHandler):
    @property
    def total(self, day: int) -> float:
        cols = [self.worksheet.cell(row=r, column=day+1).value for r in range(3, self.template.max_column + 1)]
        return sum(filter(None, cols))


"""
Meta-handler.
"""

class HandlerRotater:
    def __init__(self, root: Path, template: Path, doctype: str) -> None:
        self.to_name = get_toname(doctype, root)
        self.template = load_template(template, reduce(doctype))
        self.type = reduce(doctype, AO, FTR, InvoiceHandler)

        self.handlers = {}

    def _get_workbook(self, path: Path) -> Workbook:
        try:
            wb = load_workbook(path)
            return wb
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)
            return wb

    def get_handler(self, year: int, fund: str) -> SheetHandler:
        path = self.to_name(year, fund)
        if year not in self.handlers:
            self.handlers[year] = self.type(path=path, workbook=self._get_workbook(path), template=self.template)
        return self.handlers[year]

    def write(self) -> None:
        for handler in self.handlers.values():
            handler.write()