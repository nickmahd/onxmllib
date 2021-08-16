#!/usr/bin/env python3

from pickle import dump
from openpyxl import load_workbook

t = {
    'invoice': load_workbook('invoice.xlsx'),
    'settlement': load_workbook('settlement.xlsx')
}
with open('template.pkl', 'wb') as f: dump(t, f)